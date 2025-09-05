from flask import render_template, request, redirect, url_for, flash, make_response, jsonify
from app import app, db
from models import Course, Faculty, Room, TimeSlot, Schedule, TimetableGeneration
from scheduler import TimetableScheduler
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

@app.template_filter('time12')
def time12_filter(time_str):
    """Convert 24-hour time format to 12-hour format with AM/PM"""
    try:
        # Parse the time string (expecting HH:MM format)
        hour, minute = map(int, time_str.split(':'))
        
        # Convert to 12-hour format
        if hour == 0:
            return f"12:{minute:02d}AM"
        elif hour < 12:
            return f"{hour}:{minute:02d}AM"
        elif hour == 12:
            return f"12:{minute:02d}PM"
        else:
            return f"{hour-12}:{minute:02d}PM"
    except:
        return time_str

@app.route('/')
def index():
    """Dashboard showing overview of data and recent timetables"""
    courses_count = Course.query.count()
    faculty_count = Faculty.query.count()
    rooms_count = Room.query.count()
    timeslots_count = TimeSlot.query.count()
    recent_timetables = TimetableGeneration.query.order_by(TimetableGeneration.created_at.desc()).limit(5).all()
    
    return render_template('index.html', 
                         courses_count=courses_count,
                         faculty_count=faculty_count,
                         rooms_count=rooms_count,
                         timeslots_count=timeslots_count,
                         recent_timetables=recent_timetables)

@app.route('/courses')
def courses():
    """Manage courses"""
    courses = Course.query.all()
    return render_template('courses.html', courses=courses)

@app.route('/courses/add', methods=['POST'])
def add_course():
    """Add a new course"""
    try:
        course = Course(
            code=request.form['code'].upper(),
            name=request.form['name'],
            hours_per_week=int(request.form['hours_per_week']),
            semester=request.form['semester'],
            department=request.form['department'],
            is_lab=bool(request.form.get('is_lab'))
        )
        db.session.add(course)
        db.session.commit()
        flash('Course added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding course: {str(e)}', 'error')
    
    return redirect(url_for('courses'))

@app.route('/courses/edit/<int:course_id>', methods=['POST'])
def edit_course(course_id):
    """Edit a course"""
    try:
        course = Course.query.get_or_404(course_id)
        course.code = request.form['code'].upper()
        course.name = request.form['name']
        course.hours_per_week = int(request.form['hours_per_week'])
        course.semester = request.form['semester']
        course.department = request.form['department']
        course.is_lab = 'is_lab' in request.form
        db.session.commit()
        flash('Course updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating course: {str(e)}', 'error')
    
    return redirect(url_for('courses'))

@app.route('/courses/delete/<int:course_id>')
def delete_course(course_id):
    """Delete a course"""
    try:
        course = Course.query.get_or_404(course_id)
        # First delete all schedules that use this course
        Schedule.query.filter_by(course_id=course_id).delete()
        # Then delete the course
        db.session.delete(course)
        db.session.commit()
        flash('Course deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting course: {str(e)}', 'error')
    
    return redirect(url_for('courses'))

@app.route('/faculty')
def faculty():
    """Manage faculty"""
    faculty_list = Faculty.query.all()
    courses = Course.query.all()
    return render_template('faculty.html', faculty_list=faculty_list, courses=courses)

@app.route('/faculty/add', methods=['POST'])
def add_faculty():
    """Add a new faculty member"""
    try:
        faculty = Faculty(
            name=request.form['name'],
            email=request.form['email'],
            department=request.form['department'],
            subjects=request.form['subjects']
        )
        db.session.add(faculty)
        db.session.commit()
        flash('Faculty added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding faculty: {str(e)}', 'error')
    
    return redirect(url_for('faculty'))

@app.route('/faculty/edit/<int:faculty_id>', methods=['POST'])
def edit_faculty(faculty_id):
    """Edit a faculty member"""
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        faculty.name = request.form['name']
        faculty.email = request.form['email']
        faculty.department = request.form['department']
        faculty.subjects = request.form['subjects']
        db.session.commit()
        flash('Faculty updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating faculty: {str(e)}', 'error')
    
    return redirect(url_for('faculty'))

@app.route('/faculty/delete/<int:faculty_id>')
def delete_faculty(faculty_id):
    """Delete a faculty member"""
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        # First delete all schedules that use this faculty
        Schedule.query.filter_by(faculty_id=faculty_id).delete()
        # Then delete the faculty
        db.session.delete(faculty)
        db.session.commit()
        flash('Faculty deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting faculty: {str(e)}', 'error')
    
    return redirect(url_for('faculty'))

@app.route('/rooms')
def rooms():
    """Manage rooms"""
    rooms_list = Room.query.all()
    return render_template('rooms.html', rooms=rooms_list)

@app.route('/rooms/add', methods=['POST'])
def add_room():
    """Add a new room"""
    try:
        room = Room(
            number=request.form['number'],
            capacity=int(request.form['capacity']),
            room_type=request.form['room_type'],
            building=request.form.get('building', '')
        )
        db.session.add(room)
        db.session.commit()
        flash('Room added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding room: {str(e)}', 'error')
    
    return redirect(url_for('rooms'))

@app.route('/rooms/edit/<int:room_id>', methods=['POST'])
def edit_room(room_id):
    """Edit a room"""
    try:
        room = Room.query.get_or_404(room_id)
        room.number = request.form['number']
        room.room_type = request.form['room_type']
        room.capacity = int(request.form['capacity'])
        room.building = request.form['building'] if request.form['building'] else None
        db.session.commit()
        flash('Room updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating room: {str(e)}', 'error')
    
    return redirect(url_for('rooms'))

@app.route('/rooms/delete/<int:room_id>')
def delete_room(room_id):
    """Delete a room"""
    try:
        room = Room.query.get_or_404(room_id)
        # First delete all schedules that use this room
        Schedule.query.filter_by(room_id=room_id).delete()
        # Then delete the room
        db.session.delete(room)
        db.session.commit()
        flash('Room deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting room: {str(e)}', 'error')
    
    return redirect(url_for('rooms'))

@app.route('/generate')
def generate():
    """Generate timetable interface"""
    # Initialize default time slots if none exist
    if TimeSlot.query.count() == 0:
        default_slots = [
            # Monday - Your exact timing schedule
            ('Monday', '09:00', '09:50', 1),    # 9:00AM TO 9:50AM
            ('Monday', '09:50', '10:35', 2),   # 9:50AM TO 10:35AM
            ('Monday', '10:50', '11:35', 3),   # 10:50AM TO 11:35AM
            ('Monday', '11:35', '12:20', 4),   # 11:35AM TO 12:20PM
            ('Monday', '12:20', '13:05', 5),   # 12:20PM TO 1:05PM
            ('Monday', '14:00', '14:45', 6),   # 2:00PM TO 2:45PM
            ('Monday', '14:45', '15:30', 7),   # 2:45PM TO 3:30PM
            ('Monday', '15:45', '16:30', 8),   # 3:45PM TO 4:30PM
            # Tuesday
            ('Tuesday', '09:00', '09:50', 1),
            ('Tuesday', '09:50', '10:35', 2),
            ('Tuesday', '10:50', '11:35', 3),
            ('Tuesday', '11:35', '12:20', 4),
            ('Tuesday', '12:20', '13:05', 5),
            ('Tuesday', '14:00', '14:45', 6),
            ('Tuesday', '14:45', '15:30', 7),
            ('Tuesday', '15:45', '16:30', 8),
            # Wednesday
            ('Wednesday', '09:00', '09:50', 1),
            ('Wednesday', '09:50', '10:35', 2),
            ('Wednesday', '10:50', '11:35', 3),
            ('Wednesday', '11:35', '12:20', 4),
            ('Wednesday', '12:20', '13:05', 5),
            ('Wednesday', '14:00', '14:45', 6),
            ('Wednesday', '14:45', '15:30', 7),
            ('Wednesday', '15:45', '16:30', 8),
            # Thursday
            ('Thursday', '09:00', '09:50', 1),
            ('Thursday', '09:50', '10:35', 2),
            ('Thursday', '10:50', '11:35', 3),
            ('Thursday', '11:35', '12:20', 4),
            ('Thursday', '12:20', '13:05', 5),
            ('Thursday', '14:00', '14:45', 6),
            ('Thursday', '14:45', '15:30', 7),
            ('Thursday', '15:45', '16:30', 8),
            # Friday
            ('Friday', '09:00', '09:50', 1),
            ('Friday', '09:50', '10:35', 2),
            ('Friday', '10:50', '11:35', 3),
            ('Friday', '11:35', '12:20', 4),
            ('Friday', '12:20', '13:05', 5),
            ('Friday', '14:00', '14:45', 6),
            ('Friday', '14:45', '15:30', 7),
            ('Friday', '15:45', '16:30', 8),
        ]
        
        for day, start, end, period in default_slots:
            slot = TimeSlot(day=day, start_time=start, end_time=end, period_number=period)
            db.session.add(slot)
        db.session.commit()
    
    departments = db.session.query(Course.department).distinct().all()
    semesters = db.session.query(Course.semester).distinct().all()
    
    return render_template('generate.html', 
                         departments=[d[0] for d in departments],
                         semesters=[s[0] for s in semesters])

@app.route('/generate/run', methods=['POST'])
def run_generation():
    """Run the timetable generation algorithm"""
    try:
        department = request.form['department']
        semester = request.form['semester']
        batch = request.form['batch']
        name = f"{department} - {semester} - {batch}"
        
        # Clear existing schedules for this combination
        Schedule.query.filter_by(batch=name).delete()
        
        # Create timetable generation record
        generation = TimetableGeneration(
            name=name,
            department=department,
            semester=semester
        )
        db.session.add(generation)
        db.session.commit()
        
        # Run the scheduling algorithm
        scheduler = TimetableScheduler()
        success = scheduler.generate_timetable(department, semester, name)
        
        if success:
            generation.status = 'generated'
            flash('Timetable generated successfully!', 'success')
        else:
            generation.status = 'failed'
            flash('Failed to generate conflict-free timetable. Please check constraints.', 'error')
        
        db.session.commit()
        
        return redirect(url_for('view_timetable', generation_id=generation.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error generating timetable: {str(e)}', 'error')
        return redirect(url_for('generate'))

@app.route('/timetable/<int:generation_id>')
def view_timetable(generation_id):
    """View generated timetable"""
    generation = TimetableGeneration.query.get_or_404(generation_id)
    
    # Get all schedules for this generation (using batch as identifier)
    schedules = Schedule.query.filter_by(batch=generation.name).all()
    
    # Organize schedules by day and time
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    time_slots = TimeSlot.query.order_by(TimeSlot.period_number).all()
    
    # Create timetable grid
    timetable_grid = {}
    for day in days:
        timetable_grid[day] = {}
        for slot in time_slots:
            if slot.day == day:
                timetable_grid[day][slot.period_number] = None
    
    # Fill in the schedules
    for schedule in schedules:
        day = schedule.timeslot.day
        period = schedule.timeslot.period_number
        if day in timetable_grid and period in timetable_grid[day]:
            timetable_grid[day][period] = schedule
    
    # Faculty-wise timetable
    faculty_timetable = {}
    for schedule in schedules:
        faculty_name = schedule.faculty.name
        if faculty_name not in faculty_timetable:
            faculty_timetable[faculty_name] = {}
            for day in days:
                faculty_timetable[faculty_name][day] = {}
                for slot in time_slots:
                    if slot.day == day:
                        faculty_timetable[faculty_name][day][slot.period_number] = None
        
        day = schedule.timeslot.day
        period = schedule.timeslot.period_number
        faculty_timetable[faculty_name][day][period] = schedule
    
    # Room-wise timetable
    room_timetable = {}
    for schedule in schedules:
        room_number = schedule.room.number
        if room_number not in room_timetable:
            room_timetable[room_number] = {}
            for day in days:
                room_timetable[room_number][day] = {}
                for slot in time_slots:
                    if slot.day == day:
                        room_timetable[room_number][day][slot.period_number] = None
        
        day = schedule.timeslot.day
        period = schedule.timeslot.period_number
        room_timetable[room_number][day][period] = schedule
    
    return render_template('timetable.html',
                         generation=generation,
                         schedules=schedules,
                         timetable_grid=timetable_grid,
                         faculty_timetable=faculty_timetable,
                         room_timetable=room_timetable,
                         time_slots=time_slots,
                         days=days)

@app.route('/export/<int:generation_id>')
def export_timetable(generation_id):
    """Export timetable to Excel"""
    generation = TimetableGeneration.query.get_or_404(generation_id)
    schedules = Schedule.query.filter_by(batch=generation.name).all()
    
    # Create workbook
    wb = Workbook()
    
    # Student timetable sheet
    ws_student = wb.active
    ws_student.title = "Student Timetable"
    
    # Faculty timetable sheet
    ws_faculty = wb.create_sheet("Faculty Timetable")
    
    # Room timetable sheet
    ws_room = wb.create_sheet("Room Timetable")
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    time_slots = TimeSlot.query.order_by(TimeSlot.period_number).all()
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Student timetable
    ws_student.cell(1, 1, "Time/Day").font = header_font
    ws_student.cell(1, 1).fill = header_fill
    
    for col, day in enumerate(days, 2):
        ws_student.cell(1, col, day).font = header_font
        ws_student.cell(1, col).fill = header_fill
    
    row = 2
    for slot in time_slots:
        if slot.day == 'Monday':  # Use Monday slots as time reference
            time_str = f"{slot.start_time}-{slot.end_time}"
            ws_student.cell(row, 1, time_str)
            
            for col, day in enumerate(days, 2):
                schedule = None
                for s in schedules:
                    if s.timeslot.day == day and s.timeslot.period_number == slot.period_number:
                        schedule = s
                        break
                
                if schedule:
                    cell_text = f"{schedule.course.code}\n{schedule.faculty.name}\n{schedule.room.number}"
                    ws_student.cell(row, col, cell_text)
                    ws_student.cell(row, col).alignment = Alignment(wrap_text=True)
            
            row += 1
    
    # Create output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=timetable_{generation.name.replace(" ", "_")}.xlsx'
    
    return response
